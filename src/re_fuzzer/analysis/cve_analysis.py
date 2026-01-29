"""
CVE Analysis for ICSE Paper: Regex Engine Security Vulnerabilities

This script analyzes CVE data collected from NIST NVD to generate:
- Statistical summaries for the paper
- Visualizations (bar chart by engine, pie chart by vulnerability category)
- LaTeX tables for inclusion in the paper

Usage:
    python src/re_fuzzer/analysis/cve_analysis.py
"""

from collections import Counter, defaultdict
from pathlib import Path

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
import openpyxl

# Configure matplotlib for publication quality
plt.rcParams.update({
    'font.family': 'serif',
    'font.size': 9,
    'axes.labelsize': 9,
    'axes.titlesize': 10,
    'xtick.labelsize': 8,
    'ytick.labelsize': 8,
    'legend.fontsize': 8,
    'figure.dpi': 300,
    'savefig.dpi': 300,
    'savefig.bbox': 'tight',
    'savefig.pad_inches': 0.05,
})

# Paths
DATA_DIR = Path(__file__).parent.parent / "data"
EXCEL_FILE = DATA_DIR / "regex_cve_report.xlsx"
FIGURES_DIR = DATA_DIR / "figures"
TABLES_DIR = DATA_DIR / "tables"

# Engine name mapping: Excel sheet name -> Paper display name
ENGINE_DISPLAY_NAMES = {
    "PCRE": "PCRE",
    "PCRE2": "PCRE2",
    "PHP PCRE": "PHP",
    "Python re": "Python",
    "Oniguruma": "Oniguruma",
    "Onigmo": "Onigmo",
    "RE2": "RE2",
    "Ruby Regex": "Ruby",
    "JavaScript RegExp": "JS/TS",
    "Java Pattern": "Java",
    ".NET Regex": "C#",
    "Go regexp": "Go",
    "Rust regex": "Rust",
    "Hyperscan": "Hyperscan",
    "TRE": "TRE",
    "mrab-regex": "mrab-regex",
    "regexp2": "regexp2",
    "tiny-regex-c": "tiny-regex-c",
    "sregex": "sregex",
    "fancy-regex": "fancy-regex",
    "Bash": "Bash",
    "C C++": "C/C++",
}

# Engines to exclude from analysis
EXCLUDED_ENGINES = {"Boost.Regex"}

# CWE Definitions with names
CWE_NAMES = {
    "CWE-119": "Buffer Errors",
    "CWE-125": "Out-of-bounds Read",
    "CWE-787": "Out-of-bounds Write",
    "CWE-416": "Use After Free",
    "CWE-190": "Integer Overflow",
    "CWE-476": "NULL Pointer Deref",
    "CWE-122": "Heap Buffer Overflow",
    "CWE-121": "Stack Buffer Overflow",
    "CWE-120": "Buffer Copy w/o Size Check",
    "CWE-189": "Numeric Errors",
    "CWE-415": "Double Free",
    "CWE-674": "Uncontrolled Recursion",
    "CWE-1333": "ReDoS",
    "CWE-400": "Resource Exhaustion",
    "CWE-399": "Resource Mgmt Errors",
    "CWE-94": "Code Injection",
    "CWE-269": "Privilege Escalation",
    "CWE-20": "Input Validation",
    "CWE-200": "Info Exposure",
    "CWE-79": "XSS",
}

# CWE Categorization
MEMORY_SAFETY_CWES = {
    "CWE-119", "CWE-125", "CWE-787", "CWE-416", "CWE-190", "CWE-476",
    "CWE-122", "CWE-121", "CWE-120", "CWE-189", "CWE-415", "CWE-674",
}

REDOS_CWES = {
    "CWE-1333", "CWE-400", "CWE-399",
}

CODE_INJECTION_CWES = {
    "CWE-94",  # Code Injection (common in PHP regex)
}

# Vulnerability type colors (colorblind-friendly)
CVE_TYPE_COLORS = {
    'Memory Safety': '#CC3311',   # Red
    'ReDoS': '#EE7733',           # Orange
    'Code Injection': '#0077BB',  # Blue
    'Other': '#BBBBBB',           # Gray
}

CVE_TYPES = ['Memory Safety', 'ReDoS', 'Code Injection', 'Other']

# Engine language classification (using Excel sheet names)
C_BASED_ENGINES = {
    "PCRE", "PCRE2", "Oniguruma", "Onigmo", "TRE", "Hyperscan",
    "PHP PCRE", "RE2", "tiny-regex-c", "sregex", "C C++", "Bash"
}

MANAGED_LANGUAGE_ENGINES = {
    "Python re", "Java Pattern", ".NET Regex", "Ruby Regex",
    "JavaScript RegExp", "Go regexp", "Rust regex", "mrab-regex",
    "regexp2", "fancy-regex"
}


def load_data():
    """Load CVE data from Excel file, reading directly from sheets."""
    wb = openpyxl.load_workbook(EXCEL_FILE)

    all_cves = {}  # cve_id -> {engine, severity, cwe, ...}
    engine_cves = defaultdict(list)
    summary_data = []

    # Parse all detail sheets (not Summary)
    for sheet_name in wb.sheetnames:
        if sheet_name == "Summary" or sheet_name in EXCLUDED_ENGINES:
            continue

        sheet = wb[sheet_name]
        cves_for_engine = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                cve_id = row[0]
                cve_data = {
                    "cve_id": cve_id,
                    "engine": sheet_name,
                    "severity": row[1] if row[1] else "UNKNOWN",
                    "cvss_v3": row[2],
                    "cvss_v2": row[3],
                    "published": row[4],
                    "cwe": row[5] if row[5] else "",
                    "description": row[6] if row[6] else "",
                }
                if cve_id not in all_cves:
                    all_cves[cve_id] = cve_data
                cves_for_engine.append(cve_data)
                engine_cves[sheet_name].append(cve_data)

        # Build summary data from actual sheet contents
        critical = sum(1 for c in cves_for_engine if c["severity"] == "CRITICAL")
        high = sum(1 for c in cves_for_engine if c["severity"] == "HIGH")
        medium = sum(1 for c in cves_for_engine if c["severity"] == "MEDIUM")
        low = sum(1 for c in cves_for_engine if c["severity"] == "LOW")
        unknown = len(cves_for_engine) - (critical + high + medium + low)

        summary_data.append({
            "engine": sheet_name,
            "total": len(cves_for_engine),
            "critical": critical,
            "high": high,
            "medium": medium,
            "low": low,
            "unknown": unknown
        })

    return summary_data, all_cves, engine_cves


def categorize_cwe(cwe_str):
    """Categorize a CWE string into Memory Safety, ReDoS, Code Injection, or Other."""
    if not cwe_str:
        return "Other"

    cwes = [c.strip() for c in str(cwe_str).split(",")]

    # Check for Memory Safety first (most common)
    for cwe in cwes:
        if cwe in MEMORY_SAFETY_CWES:
            return "Memory Safety"

    # Check for ReDoS
    for cwe in cwes:
        if cwe in REDOS_CWES:
            return "ReDoS"

    # Check for Code Injection
    for cwe in cwes:
        if cwe in CODE_INJECTION_CWES:
            return "Code Injection"

    return "Other"


def get_specific_cwe_category(cwe_str):
    """Get more specific CWE category for detailed pie chart."""
    if not cwe_str:
        return "Unknown/None"

    cwes = [c.strip() for c in str(cwe_str).split(",")]
    primary_cwe = cwes[0] if cwes else ""

    # Map to specific categories
    if primary_cwe in {"CWE-119", "CWE-125", "CWE-787", "CWE-122", "CWE-121", "CWE-120"}:
        return f"CWE-119/125/787: Buffer Errors"
    elif primary_cwe in {"CWE-190", "CWE-189"}:
        return f"CWE-190: Integer Overflow"
    elif primary_cwe in {"CWE-416", "CWE-415"}:
        return f"CWE-416: Use After Free"
    elif primary_cwe == "CWE-476":
        return f"CWE-476: NULL Ptr Deref"
    elif primary_cwe == "CWE-674":
        return f"CWE-674: Stack Overflow"
    elif primary_cwe == "CWE-1333":
        return f"CWE-1333: ReDoS"
    elif primary_cwe in {"CWE-400", "CWE-399"}:
        return f"CWE-400: Resource Exhaustion"
    elif primary_cwe == "CWE-94":
        return f"CWE-94: Code Injection"
    elif primary_cwe == "CWE-269":
        return f"CWE-269: Privilege Issues"
    elif primary_cwe == "CWE-20":
        return f"CWE-20: Input Validation"
    elif primary_cwe == "CWE-200":
        return f"CWE-200: Info Exposure"
    elif primary_cwe == "CWE-79":
        return f"CWE-79: XSS"
    elif primary_cwe.startswith("CWE-"):
        return f"{primary_cwe}: Other"
    else:
        return "Unknown/None"


def analyze_cwe_distribution(all_cves):
    """Analyze CWE distribution across all CVEs."""
    cwe_counts = Counter()
    category_counts = Counter()
    specific_counts = Counter()

    for cve in all_cves.values():
        cwe_str = cve.get("cwe", "")
        category = categorize_cwe(cwe_str)
        category_counts[category] += 1

        specific = get_specific_cwe_category(cwe_str)
        specific_counts[specific] += 1

        if cwe_str:
            for cwe in str(cwe_str).split(","):
                cwe = cwe.strip()
                if cwe.startswith("CWE"):
                    cwe_counts[cwe] += 1

    return cwe_counts, category_counts, specific_counts


def analyze_cwe_by_engine(engine_cves):
    """Analyze CWE categories per engine for stacked bar chart."""
    engine_cwe_types = {}

    for engine, cves in engine_cves.items():
        type_counts = Counter()
        for cve in cves:
            category = categorize_cwe(cve.get("cwe", ""))
            type_counts[category] += 1
        engine_cwe_types[engine] = type_counts

    return engine_cwe_types


def analyze_severity(summary_data):
    """Analyze severity distribution."""
    totals = {
        "critical": 0,
        "high": 0,
        "medium": 0,
        "low": 0,
        "unknown": 0
    }

    for entry in summary_data:
        totals["critical"] += entry["critical"]
        totals["high"] += entry["high"]
        totals["medium"] += entry["medium"]
        totals["low"] += entry["low"]
        totals["unknown"] += entry["unknown"]

    return totals


def analyze_language_correlation(summary_data):
    """Analyze CVE counts by implementation language."""
    c_based_total = 0
    managed_total = 0

    for entry in summary_data:
        if entry["engine"] in C_BASED_ENGINES:
            c_based_total += entry["total"]
        elif entry["engine"] in MANAGED_LANGUAGE_ENGINES:
            managed_total += entry["total"]

    return c_based_total, managed_total


def create_stacked_bar_chart(summary_data, engine_cwe_types, output_path):
    """Create stacked horizontal bar chart of CVE types by engine (similar to bug_types_by_engine)."""
    # Include all engines, sort ascending (largest will be at top in horizontal bar)
    data = [(e["engine"], e["total"]) for e in summary_data]
    data.sort(key=lambda x: x[1])  # Ascending order

    engines = [d[0] for d in data]
    display_names = [ENGINE_DISPLAY_NAMES.get(e, e) for e in engines]

    fig, ax = plt.subplots(figsize=(3.5, 3.5))

    y_pos = np.arange(len(engines))

    # Build stacked bars
    left = np.zeros(len(engines))
    for cve_type in CVE_TYPES:
        values = []
        for engine in engines:
            count = engine_cwe_types.get(engine, {}).get(cve_type, 0)
            values.append(count)
        values = np.array(values)
        ax.barh(y_pos, values, left=left, label=cve_type,
                color=CVE_TYPE_COLORS[cve_type], height=0.7)
        left += values

    ax.set_yticks(y_pos)
    ax.set_yticklabels(display_names)
    ax.set_xlabel('Number of CVEs')

    # Set x-axis limit with some padding
    max_total = max(e["total"] for e in summary_data if e["total"] > 0)
    ax.set_xlim(0, max_total * 1.1)

    # Legend below plot
    ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.12),
              ncol=2, frameon=False, columnspacing=0.8)

    plt.tight_layout()
    plt.savefig(output_path, format='pdf', bbox_inches='tight')
    plt.savefig(output_path.with_suffix('.png'), format='png', bbox_inches='tight')
    plt.close()
    print(f"Saved: {output_path}")


def create_bar_chart(summary_data, output_path):
    """Create bar chart of CVEs by engine."""
    # Include all engines, sort by total count (ascending for bottom-to-top)
    data = [(e["engine"], e["total"]) for e in summary_data]
    data.sort(key=lambda x: x[1])

    engines = [d[0] for d in data]
    display_names = [ENGINE_DISPLAY_NAMES.get(e, e) for e in engines]
    counts = [d[1] for d in data]

    # Color by engine type
    colors = []
    for engine in engines:
        if engine in C_BASED_ENGINES:
            colors.append("#CC3311")  # Red for C-based
        elif engine in MANAGED_LANGUAGE_ENGINES:
            colors.append("#0077BB")  # Blue for managed
        else:
            colors.append("#BBBBBB")  # Gray for other

    fig, ax = plt.subplots(figsize=(3.5, 3.5))
    bars = ax.barh(range(len(engines)), counts, color=colors, height=0.7)
    ax.set_yticks(range(len(engines)))
    ax.set_yticklabels(display_names)
    ax.set_xlabel("Number of CVEs")

    # Add count labels
    for bar, count in zip(bars, counts):
        ax.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height()/2,
                str(count), va='center', fontsize=7)

    # Legend
    c_patch = mpatches.Patch(color="#CC3311", label="C/C++ based")
    managed_patch = mpatches.Patch(color="#0077BB", label="Managed language")
    ax.legend(handles=[c_patch, managed_patch], loc='lower right', fontsize=7)

    plt.tight_layout()
    plt.savefig(output_path, bbox_inches="tight", dpi=300)
    plt.savefig(output_path.with_suffix(".png"), bbox_inches="tight", dpi=300)
    plt.close()
    print(f"Saved: {output_path}")


def create_pie_chart(specific_counts, output_path):
    """Create pie chart with specific CWE categories."""
    # Group and sort by count
    counts = dict(specific_counts.most_common())

    # Merge small categories into "Other"
    threshold = 3
    main_categories = {}
    other_count = 0

    for category, count in counts.items():
        if count >= threshold and "Other" not in category and "Unknown" not in category:
            main_categories[category] = count
        else:
            other_count += count

    if other_count > 0:
        main_categories["Other/Unknown"] = other_count

    # Sort by count descending
    sorted_cats = sorted(main_categories.items(), key=lambda x: x[1], reverse=True)
    labels = [c[0] for c in sorted_cats]
    sizes = [c[1] for c in sorted_cats]

    # Color scheme
    colors = plt.cm.Set3(np.linspace(0, 1, len(labels)))

    fig, ax = plt.subplots(figsize=(5, 4))

    def autopct_func(pct):
        count = int(round(pct / 100 * sum(sizes)))
        return f'{pct:.0f}%\n({count})'

    wedges, texts, autotexts = ax.pie(
        sizes, labels=None, autopct=autopct_func,
        colors=colors, startangle=90, pctdistance=0.75,
        textprops={'fontsize': 8}
    )

    # Legend with labels
    ax.legend(wedges, labels, loc='center left', bbox_to_anchor=(1, 0.5),
              fontsize=8, frameon=False)

    plt.tight_layout()
    plt.savefig(output_path, bbox_inches="tight", dpi=300)
    plt.savefig(output_path.with_suffix(".png"), bbox_inches="tight", dpi=300)
    plt.close()
    print(f"Saved: {output_path}")


def create_latex_table(summary_data, output_path):
    """Create LaTeX table for the paper."""
    # Filter and sort
    data = [e for e in summary_data if e["total"] > 0]
    data.sort(key=lambda x: x["total"], reverse=True)

    lines = [
        r"\begin{table}[t]",
        r"\centering",
        r"\caption{CVEs by regex engine (2005--2025). Source: NIST NVD.}",
        r"\label{tab:cve_summary}",
        r"\begin{tabular}{lrrrrr}",
        r"\toprule",
        r"Engine & Total & Critical & High & Medium & Low \\",
        r"\midrule",
    ]

    for entry in data:
        display_name = ENGINE_DISPLAY_NAMES.get(entry["engine"], entry["engine"])
        engine = display_name.replace("_", r"\_").replace("#", r"\#")
        line = f"{engine} & {entry['total']} & {entry['critical'] or '--'} & {entry['high'] or '--'} & {entry['medium'] or '--'} & {entry['low'] or '--'} \\\\"
        lines.append(line)

    # Totals
    total_cves = sum(e["total"] for e in data)
    total_critical = sum(e["critical"] for e in data)
    total_high = sum(e["high"] for e in data)
    total_medium = sum(e["medium"] for e in data)
    total_low = sum(e["low"] for e in data)

    lines.extend([
        r"\midrule",
        f"\\textbf{{Total}} & \\textbf{{{total_cves}}} & \\textbf{{{total_critical}}} & \\textbf{{{total_high}}} & \\textbf{{{total_medium}}} & \\textbf{{{total_low}}} \\\\",
        r"\bottomrule",
        r"\end{tabular}",
        r"\end{table}",
    ])

    with open(output_path, "w") as f:
        f.write("\n".join(lines))
    print(f"Saved: {output_path}")


def print_statistics(summary_data, all_cves, cwe_counts, category_counts, severity_totals):
    """Print statistics for use in paper prose."""
    print("\n" + "=" * 70)
    print("CVE ANALYSIS STATISTICS FOR ICSE PAPER")
    print("=" * 70)

    # Basic counts
    total_entries = sum(e["total"] for e in summary_data)
    unique_cves = len(all_cves)
    engines_with_cves = sum(1 for e in summary_data if e["total"] > 0)

    print(f"\n[SCALE]")
    print(f"  Total CVE entries (with duplicates): {total_entries}")
    print(f"  Unique CVEs: {unique_cves}")
    print(f"  Engines with CVEs: {engines_with_cves}")

    # Severity
    print(f"\n[SEVERITY]")
    critical = severity_totals["critical"]
    high = severity_totals["high"]
    critical_high = critical + high
    critical_high_pct = (critical_high / unique_cves) * 100 if unique_cves else 0
    print(f"  Critical: {critical}")
    print(f"  High: {high}")
    print(f"  Critical + High: {critical_high} ({critical_high_pct:.1f}%)")
    print(f"  Medium: {severity_totals['medium']}")
    print(f"  Low: {severity_totals['low']}")
    print(f"  Unknown: {severity_totals['unknown']}")

    # Vulnerability categories
    print(f"\n[VULNERABILITY CATEGORIES]")
    total_categorized = sum(category_counts.values())
    for category in CVE_TYPES:
        count = category_counts.get(category, 0)
        pct = (count / total_categorized) * 100 if total_categorized else 0
        print(f"  {category}: {count} ({pct:.1f}%)")

    # Top CWEs
    print(f"\n[TOP 10 CWEs]")
    for cwe, count in cwe_counts.most_common(10):
        name = CWE_NAMES.get(cwe, "")
        print(f"  {cwe} ({name}): {count}")

    # Top engines
    print(f"\n[TOP ENGINES BY CVE COUNT]")
    sorted_engines = sorted(summary_data, key=lambda x: x["total"], reverse=True)
    for entry in sorted_engines[:8]:
        if entry["total"] > 0:
            display = ENGINE_DISPLAY_NAMES.get(entry["engine"], entry["engine"])
            print(f"  {display}: {entry['total']} (Critical: {entry['critical']}, High: {entry['high']})")

    # Language correlation
    c_based, managed = analyze_language_correlation(summary_data)
    print(f"\n[LANGUAGE CORRELATION]")
    print(f"  C/C++ based engines: {c_based} CVEs")
    print(f"  Managed language engines: {managed} CVEs")

    # Key findings for prose
    print("\n" + "=" * 70)
    print("KEY FINDINGS FOR PROSE")
    print("=" * 70)

    memory_pct = (category_counts.get("Memory Safety", 0) / total_categorized) * 100 if total_categorized else 0
    redos_pct = (category_counts.get("ReDoS", 0) / total_categorized) * 100 if total_categorized else 0
    injection_pct = (category_counts.get("Code Injection", 0) / total_categorized) * 100 if total_categorized else 0

    print(f"""
1. SCALE: We identified {unique_cves} unique CVEs across {engines_with_cves} regex engines
   from 2005-2025 in the NIST National Vulnerability Database.

2. SEVERITY: {critical_high_pct:.0f}% of vulnerabilities are CRITICAL or HIGH severity
   ({critical_high} of {unique_cves}), with {critical} rated CRITICAL.

3. MEMORY DOMINATES: {memory_pct:.0f}% of CVEs are memory safety issues (buffer overflows,
   out-of-bounds access, use-after-free), concentrated in C-based engines.

4. ReDoS: {redos_pct:.0f}% of CVEs relate to catastrophic backtracking (ReDoS),
   affecting both native and managed implementations.

5. Code Injection: {injection_pct:.0f}% relate to code injection (primarily PHP PCRE).

6. C ENGINE CONCENTRATION: PCRE alone accounts for {sorted_engines[0]['total']} CVEs
   ({sorted_engines[0]['total']/unique_cves*100:.0f}% of total). C-based engines have {c_based} CVEs
   vs {managed} for managed language implementations.
""")


def main():
    """Main analysis function."""
    # Ensure output directories exist
    FIGURES_DIR.mkdir(parents=True, exist_ok=True)
    TABLES_DIR.mkdir(parents=True, exist_ok=True)

    # Load data
    print("Loading CVE data from Excel...")
    summary_data, all_cves, engine_cves = load_data()

    # Analyze
    print("Analyzing CWE distribution...")
    cwe_counts, category_counts, specific_counts = analyze_cwe_distribution(all_cves)
    engine_cwe_types = analyze_cwe_by_engine(engine_cves)

    print("Analyzing severity distribution...")
    severity_totals = analyze_severity(summary_data)

    # Generate outputs
    print("\nGenerating visualizations...")
    create_bar_chart(summary_data, FIGURES_DIR / "cves_by_engine.pdf")
    create_pie_chart(specific_counts, FIGURES_DIR / "cwe_categories.pdf")
    create_stacked_bar_chart(summary_data, engine_cwe_types, FIGURES_DIR / "cve_types_by_engine.pdf")

    print("\nGenerating LaTeX table...")
    create_latex_table(summary_data, TABLES_DIR / "cve_summary.tex")

    # Print statistics
    print_statistics(summary_data, all_cves, cwe_counts, category_counts, severity_totals)


if __name__ == "__main__":
    main()
