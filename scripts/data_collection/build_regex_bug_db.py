#!/usr/bin/env python3
import argparse
import json
import os
import sys
from typing import List, Dict, Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment
import re


def extract_content_row(obj: Dict[str, Any]) -> Dict[str, Any]:
    """
    Given one NDJSON line as a dict, extract the 'content' field and
    return a dict of columns.

    Structure expected (based on your example):

    {
      "response": {
        "body": {
          "choices": [
            {
              "message": {
                "content": "{...json string...}"
              }
            }
          ]
        }
      }
    }

    We:
    - pull that content string
    - try json.loads on it
    - if that fails, just return {"content": <raw string>}
    """
    try:
        content_str = obj["response"]["body"]["choices"][0]["message"]["content"]
    except (KeyError, IndexError, TypeError):
        # Nothing usable
        return {}

    # Try to parse content as JSON
    if isinstance(content_str, str):
        try:
            parsed = json.loads(content_str)
            if isinstance(parsed, dict):
                return parsed
            else:
                # If it's not a dict, put it under 'content'
                return {"content": parsed}
        except json.JSONDecodeError:
            # Not JSON, just treat as raw string
            return {"content": content_str}
    else:
        # Unexpected type, store as-is
        return {"content": content_str}


def read_ndjson(path: str) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with open(path, "r", encoding="utf-8") as f:
        for line_no, line in enumerate(f, 1):
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError as e:
                print(f"Warning: could not parse JSON on line {line_no}: {e}", file=sys.stderr)
                continue

            row = extract_content_row(obj)
            if row:
                rows.append(row)
    return rows


def ensure_workbook(path: str):
    if os.path.exists(path):
        wb = load_workbook(path)
        created_new = False
    else:
        wb = Workbook()
        created_new = True
        # Remove default sheet if we’ll create our own
        default_sheet = wb.active
        wb.remove(default_sheet)
    return wb, created_new


def normalize_cell_value(value: Any) -> Any:
    """
    Prepare a Python value for safe writing into an Excel cell.
    - Dicts/lists are JSON-encoded so openpyxl can write them
    - Bytes are decoded to UTF-8 (with replacement) first
    - Strings have Excel-illegal control characters stripped
    """
    if value is None:
        return None

    if isinstance(value, (list, dict)):
        try:
            return json.dumps(value, ensure_ascii=False)
        except Exception:
            return str(value)

    if isinstance(value, bytes):
        try:
            value = value.decode("utf-8", errors="replace")
        except Exception:
            value = str(value)

    if isinstance(value, str):
        # Remove characters that Excel cannot handle
        return ILLEGAL_CHARACTERS_RE.sub("", value)

    # ints, floats, bools are fine as-is
    return value


def add_sheet_with_table(
    wb,
    sheet_name: str,
    rows: List[Dict[str, Any]],
):
    if not rows:
        raise ValueError("No rows to write; did not find any 'content' entries.")

    # Collect all keys for columns, preserving order of first row
    columns = list(rows[0].keys())
    for r in rows[1:]:
        for k in r.keys():
            if k not in columns:
                columns.append(k)

    # If sheet already exists, delete it and recreate
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)

    # Write header
    ws.append(columns)

    # Write data rows (with minimal preview-friendly coercions)
    for r in rows:
        row_values = []
        for col in columns:
            value = normalize_cell_value(r.get(col, None))
            # Escape Excel formula interpretation for 'log' field when it starts with '='
            if col == "log" and isinstance(value, str) and value.startswith("="):
                value = "\\" + value
            row_values.append(value)
        ws.append(row_values)

    # Create an Excel "Table" for nicer formatting/filtering
    last_row = len(rows) + 1  # +1 for header
    last_col = len(columns)
    last_col_letter = get_column_letter(last_col)
    ref = f"A1:{last_col_letter}{last_row}"

    # Excel table display names must be unique and alphanumeric-ish
    display_name = re.sub(r"[^A-Za-z0-9_]", "_", sheet_name)
    if not display_name:
        display_name = "Table1"

    table = Table(displayName=display_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Preview-oriented display: narrow columns, no wrapping, fixed row height
    # - Users can click into a cell to view full content
    default_col_width = 18  # small, reasonable preview width
    for col_idx in range(1, last_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = default_col_width

    # Disable wrapping and enable shrink-to-fit for all cells
    preview_alignment = Alignment(wrap_text=False, shrink_to_fit=True)
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.alignment = preview_alignment

    # Keep row height compact to avoid tall rows from embedded newlines
    compact_height = 15  # Excel default-ish
    for row_idx in range(1, last_row + 1):
        ws.row_dimensions[row_idx].height = compact_height


def main():
    parser = argparse.ArgumentParser(
        description="Convert NDJSON bug report analyses (content field) to an Excel sheet."
    )
    parser.add_argument("ndjson", help="Input NDJSON file path")
    parser.add_argument("xlsx", help="Output .xlsx file path (existing or new)")
    parser.add_argument(
        "--sheet-name",
        "-s",
        required=True,
        help="Name of the sheet/tab to create inside the workbook",
    )

    args = parser.parse_args()

    rows = read_ndjson(args.ndjson)

    # Remove non-bug rows
    processed_rows = []

    for r in rows:
        if (
            r.get("is_real_bug") == False
        ):
            continue
        else:
            processed_rows.append(r)

    rows = processed_rows

    if not rows:
        print("No usable rows found in the NDJSON (no 'content' fields).", file=sys.stderr)
        sys.exit(1)

    wb, _ = ensure_workbook(args.xlsx)
    try:
        add_sheet_with_table(wb, args.sheet_name, rows)
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

    wb.save(args.xlsx)
    print(f"Wrote {len(rows)} rows to '{args.xlsx}' in sheet '{args.sheet_name}'.")


if __name__ == "__main__":
    main()
