import requests
import time
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

NVD_API_BASE = "https://services.nvd.nist.gov/rest/json/cves/2.0"
MAX_RETRIES = 3
BASE_DELAY = 7

# Comprehensive regex engine definitions with improved filtering
REGEX_ENGINES = {
    "PCRE": {
        "keywords": ["pcre"],
        "must_contain_any": [
            ["libpcre1"],
            ["pcre", "8."],  # PCRE version 8.x
            ["pcre"],
        ],
        "exclude_contains": ["pcre2", "pcre-2"],
        "allow_shared": True,  # Can share CVEs with PCRE2
    },
    "PCRE2": {
        "keywords": ["pcre2"],
        "must_contain_any": [
            ["pcre2"],
            ["libpcre2"],
            ["pcre", "10."],  # PCRE2 version 10.x
        ],
        "exclude_contains": [],
        "allow_shared": True,  # Can share CVEs with PCRE
    },
    "Python re": {
        "keywords": ["python regex", "python regular expression"],
        "must_contain_any": [
            ["python", "regex"],
            ["python", "regular expression"],
            ["python", "re module"]
        ],
        "exclude_contains": [],
    },
    "RE2": {
        "keywords": ["re2"],
        "must_contain_any": [
            ["google re2"],
            [" re2 "],
            ["re2,"],
            ["re2."],
        ],
        "exclude_contains": ["re2c", "libre2"],
    },
    "Oniguruma": {
        "keywords": ["oniguruma"],
        "must_contain": ["oniguruma"],
        "exclude_contains": ["onigmo"],
    },
    "Onigmo": {
        "keywords": ["onigmo"],
        "must_contain": ["onigmo"],
        "exclude_contains": [],
    },
    "Java Pattern": {
        "keywords": ["java.util.regex", "java regex"],
        "must_contain_any": [
            ["java.util.regex"],
            ["java", "regex"],
            ["java", "regular expression"]
        ],
        "exclude_contains": ["javascript"],
    },
    ".NET Regex": {
        "keywords": [".net regex", "system.text.regularexpressions"],
        "must_contain_any": [
            ["system.text.regularexpressions"],
            [".net", "regex"],
            ["dotnet", "regex"],
            [".net framework", "regex"]
        ],
        "exclude_contains": [],
    },
    "JavaScript RegExp": {
        "keywords": ["javascript regexp", "v8 regexp", "v8 regex"],
        "must_contain_any": [
            ["javascript", "regexp"],
            ["javascript", "regex"],
            ["v8", "regexp"],
            ["v8", "regex"],
            ["ecmascript", "regexp"]
        ],
        # Exclude if it's clearly about PCRE
        "exclude_contains": ["libpcre", "pcre before", "pcre through", "pcre_"],
    },
    "Ruby Regex": {
        "keywords": ["ruby regex", "ruby regexp"],
        "must_contain_any": [
            ["ruby", "regex"],
            ["ruby", "regexp"],
            ["ruby", "regular expression"]
        ],
        "exclude_contains": [],
    },
    "PHP PCRE": {
        "keywords": ["php pcre", "php preg", "php regex"],
        "must_contain_any": [
            ["php", "pcre"],
            ["php", "preg_"],
            ["php", "preg "],
        ],
        "exclude_contains": [],
    },
    "Go regexp": {
        "keywords": ["go regexp", "golang regexp"],
        "must_contain_any": [
            ["regexp.compile", "go"],
            ["regexp", "golang"],
            ["go before", "regexp"],
            ["go 1.", "regexp"],
        ],
        "exclude_contains": ["google", "mongodb", "rego"],
    },
    "Boost.Regex": {
        "keywords": ["boost regex"],
        "must_contain_any": [
            ["boost", "regex"],
            ["boost::regex"]
        ],
        "exclude_contains": [],
    },
    "C++ <regex>": {
        "keywords": ["c++ regex", "std::regex"],
        "must_contain_any": [
            ["std::regex"],
            ["c++", "regex"],
            ["c++11", "regex"],
        ],
        "exclude_contains": [],
    },
    "Rust regex": {
        "keywords": ["rust regex"],
        "must_contain_any": [
            ["rust", "regex"],
            ["rust-lang", "regex"],
        ],
        "exclude_contains": [],
    },
    "Bash glibc regex": {
        "keywords": ["bash regex", "glibc regex"],
        "must_contain_any": [
            ["bash", "regex"],
            ["bash", "regular expression"],
            ["glibc", "regex"],
            ["gnu libc", "regex"],
        ],
        "exclude_contains": [],
    },
    "Hyperscan": {
        "keywords": ["hyperscan"],
        "must_contain": ["hyperscan"],
        "exclude_contains": [],
    },
    "TRE": {
        "keywords": ["tre regex", "tre library"],
        "must_contain_any": [
            ["tre regex"],
            ["tre library"],
            ["libtre"],
        ],
        "exclude_contains": [],
    },
    "regexp2": {
        "keywords": ["regexp2"],
        "must_contain": ["regexp2"],
        "exclude_contains": [],
    },
    "fancy-regex": {
        "keywords": ["fancy-regex", "fancy regex"],
        "must_contain_any": [
            ["fancy-regex"],
            ["fancy regex"],
        ],
        "exclude_contains": [],
    },
    "sregex": {
        "keywords": ["sregex"],
        "must_contain": ["sregex"],
        "exclude_contains": [],
    },
    "tiny-regex-c": {
        "keywords": ["tiny-regex-c", "tiny regex"],
        "must_contain_any": [
            ["tiny-regex-c"],
            ["tiny regex"],
        ],
        "exclude_contains": [],
    },
}

def search_nvd_with_retry(params, description):
    """Search NVD API with retry logic"""
    for attempt in range(MAX_RETRIES):
        try:
            response = requests.get(NVD_API_BASE, params=params, timeout=30)
            response.raise_for_status()
            data = response.json()
            vulnerabilities = data.get("vulnerabilities", [])
            total = data.get("totalResults", 0)
            print(f"  ✓ {description}: {len(vulnerabilities)} CVEs (total: {total})")
            return vulnerabilities
        except requests.exceptions.HTTPError as e:
            if e.response.status_code in [503, 429]:
                wait_time = BASE_DELAY * (2 ** attempt)
                print(f"  ⚠ Rate limited, waiting {wait_time}s... (attempt {attempt + 1}/{MAX_RETRIES})")
                time.sleep(wait_time)
            else:
                print(f"  ✗ HTTP Error - {description}: {e}")
                return []
        except requests.exceptions.RequestException as e:
            print(f"  ✗ Error - {description}: {e}")
            if attempt < MAX_RETRIES - 1:
                time.sleep(BASE_DELAY)
            else:
                return []

    print(f"  ✗ Max retries exceeded for {description}")
    return []

def get_cves_by_keyword(keyword):
    """Search CVEs by keyword with pagination"""
    all_cves = []
    start_index = 0
    results_per_page = 2000

    while True:
        params = {
            "keywordSearch": keyword,
            "resultsPerPage": results_per_page,
            "startIndex": start_index
        }

        cves = search_nvd_with_retry(params, f"'{keyword}' (start: {start_index})")
        if not cves:
            break

        all_cves.extend(cves)

        if len(cves) < results_per_page:
            break

        start_index += results_per_page
        time.sleep(BASE_DELAY)

    return all_cves

def matches_engine(cve_description, engine_config):
    """Check if CVE description matches engine with flexible pattern matching"""
    if not cve_description:
        return False

    desc_lower = cve_description.lower()

    # Check exclude patterns first
    for exclude_term in engine_config.get("exclude_contains", []):
        if exclude_term.lower() in desc_lower:
            return False

    # Check must_contain (all terms must be present)
    if "must_contain" in engine_config:
        for term in engine_config["must_contain"]:
            if term.lower() not in desc_lower:
                return False
        return True

    # Check must_contain_any (at least one set of terms must all be present)
    if "must_contain_any" in engine_config:
        for term_group in engine_config["must_contain_any"]:
            if all(term.lower() in desc_lower for term in term_group):
                return True
        return False

    return False

def extract_cve_data(cve_item):
    """Extract relevant data from CVE item"""
    cve = cve_item.get("cve", {})
    cve_id = cve.get("id", "")

    # Get description
    descriptions = cve.get("descriptions", [])
    description = ""
    for desc in descriptions:
        if desc.get("lang") == "en":
            description = desc.get("value", "")
            break

    # Get dates
    published = cve.get("published", "")[:10] if cve.get("published") else ""
    last_modified = cve.get("lastModified", "")[:10] if cve.get("lastModified") else ""

    # Get CVSS scores
    cvss_v3_score = ""
    cvss_v3_severity = ""
    cvss_v2_score = ""

    metrics = cve.get("metrics", {})

    # CVSS v3.x
    for version in ["cvssMetricV31", "cvssMetricV30"]:
        if version in metrics and metrics[version]:
            cvss_data = metrics[version][0].get("cvssData", {})
            cvss_v3_score = cvss_data.get("baseScore", "")
            cvss_v3_severity = cvss_data.get("baseSeverity", "")
            break

    # CVSS v2
    if "cvssMetricV2" in metrics and metrics["cvssMetricV2"]:
        cvss_data = metrics["cvssMetricV2"][0].get("cvssData", {})
        cvss_v2_score = cvss_data.get("baseScore", "")

    # Get references
    references = cve.get("references", [])
    ref_urls = [ref.get("url", "") for ref in references[:5]]

    # Get CWE
    weaknesses = cve.get("weaknesses", [])
    cwe_ids = []
    for weakness in weaknesses:
        for desc in weakness.get("description", []):
            if desc.get("lang") == "en":
                value = desc.get("value", "")
                if value and value not in cwe_ids:
                    cwe_ids.append(value)

    return {
        "CVE ID": cve_id,
        "Description": description,
        "Published Date": published,
        "Last Modified": last_modified,
        "CVSS v3 Score": cvss_v3_score,
        "CVSS v3 Severity": cvss_v3_severity,
        "CVSS v2 Score": cvss_v2_score,
        "CWE": ", ".join(cwe_ids),
        "References": "\n".join(ref_urls)
    }

def check_pcre_shared(description):
    """Check if a CVE affects both PCRE and PCRE2"""
    desc_lower = description.lower()
    has_pcre1 = any(term in desc_lower for term in ["libpcre1", "pcre 8."])
    has_pcre2 = any(term in desc_lower for term in ["libpcre2", "pcre2", "pcre 10."])
    return has_pcre1 and has_pcre2

def collect_cves_for_engine(engine_name, engine_config):
    """Collect all CVEs for a specific engine"""
    print(f"\n{'='*70}")
    print(f"Processing: {engine_name}")
    print(f"{'='*70}")

    all_cve_items = {}

    # Search by keywords
    for keyword in engine_config.get("keywords", []):
        print(f"\nSearching by keyword: {keyword}")
        cves = get_cves_by_keyword(keyword)

        for cve_item in cves:
            cve_id = cve_item.get("cve", {}).get("id", "")
            if cve_id and cve_id not in all_cve_items:
                cve_data = extract_cve_data(cve_item)
                # Filter by description pattern
                if matches_engine(cve_data["Description"], engine_config):
                    all_cve_items[cve_id] = cve_data

        time.sleep(BASE_DELAY)

    # Convert to list and sort by CVSS score
    cve_list = list(all_cve_items.values())
    cve_list.sort(key=lambda x: (
        float(x.get("CVSS v3 Score") or x.get("CVSS v2 Score") or 0),
        x.get("CVE ID", "")
    ), reverse=True)

    print(f"\n✓ Total CVEs for {engine_name}: {len(cve_list)}")
    return cve_list

def add_shared_pcre_cves(cve_data_by_engine):
    """Add CVEs that affect both PCRE and PCRE2 to both sheets"""
    if "PCRE" not in cve_data_by_engine or "PCRE2" not in cve_data_by_engine:
        return

    pcre_cves = {cve["CVE ID"]: cve for cve in cve_data_by_engine["PCRE"]}
    pcre2_cves = {cve["CVE ID"]: cve for cve in cve_data_by_engine["PCRE2"]}

    # Check PCRE CVEs for ones that should be in PCRE2
    for cve_id, cve in pcre_cves.items():
        if cve_id not in pcre2_cves and check_pcre_shared(cve["Description"]):
            cve_data_by_engine["PCRE2"].append(cve)
            print(f"  ✓ Added {cve_id} to PCRE2 (affects both)")

    # Check PCRE2 CVEs for ones that should be in PCRE
    for cve_id, cve in pcre2_cves.items():
        if cve_id not in pcre_cves and check_pcre_shared(cve["Description"]):
            cve_data_by_engine["PCRE"].append(cve)
            print(f"  ✓ Added {cve_id} to PCRE (affects both)")

    # Re-sort both lists
    for engine in ["PCRE", "PCRE2"]:
        cve_data_by_engine[engine].sort(key=lambda x: (
            float(x.get("CVSS v3 Score") or x.get("CVSS v2 Score") or 0),
            x.get("CVE ID", "")
        ), reverse=True)

def create_excel_report(cve_data_by_engine, output_file):
    """Create comprehensive Excel report"""
    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Create summary sheet
    summary_sheet = wb.create_sheet("Summary", 0)

    # Title
    summary_sheet["A1"] = "Regex Engine CVE Security Report"
    summary_sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    summary_sheet["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    summary_sheet.merge_cells("A1:G1")
    summary_sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    summary_sheet["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}"
    summary_sheet["A2"].font = Font(italic=True, size=10)

    summary_sheet["A3"] = "Data Source: NIST National Vulnerability Database (NVD) API v2.0"
    summary_sheet["A3"].font = Font(italic=True, size=9)

    # Headers
    headers = ["Regex Engine", "Total CVEs", "Critical", "High", "Medium", "Low", "None/Unknown"]
    for col_idx, header in enumerate(headers, start=1):
        cell = summary_sheet.cell(row=5, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Populate summary
    row = 6
    for engine_name in sorted(cve_data_by_engine.keys()):
        cves = cve_data_by_engine[engine_name]

        critical = sum(1 for c in cves if c.get("CVSS v3 Severity") == "CRITICAL")
        high = sum(1 for c in cves if c.get("CVSS v3 Severity") == "HIGH")
        medium = sum(1 for c in cves if c.get("CVSS v3 Severity") == "MEDIUM")
        low = sum(1 for c in cves if c.get("CVSS v3 Severity") == "LOW")
        none_unknown = len(cves) - (critical + high + medium + low)

        summary_sheet[f"A{row}"] = engine_name
        summary_sheet[f"B{row}"] = len(cves)
        summary_sheet[f"C{row}"] = critical if critical > 0 else ""
        summary_sheet[f"D{row}"] = high if high > 0 else ""
        summary_sheet[f"E{row}"] = medium if medium > 0 else ""
        summary_sheet[f"F{row}"] = low if low > 0 else ""
        summary_sheet[f"G{row}"] = none_unknown if none_unknown > 0 else ""

        # Color code severity cells
        if critical > 0:
            summary_sheet[f"C{row}"].font = Font(color="C00000", bold=True)
        if high > 0:
            summary_sheet[f"D{row}"].font = Font(color="FF6600", bold=True)
        if medium > 0:
            summary_sheet[f"E{row}"].font = Font(color="FFC000")

        row += 1

    # Column widths
    summary_sheet.column_dimensions["A"].width = 25
    for col in ["B", "C", "D", "E", "F", "G"]:
        summary_sheet.column_dimensions[col].width = 14

    # Create detail sheets
    for engine_name in sorted(cve_data_by_engine.keys()):
        cves = cve_data_by_engine[engine_name]
        if not cves:
            continue

        sheet_name = engine_name[:31]
        sheet = wb.create_sheet(sheet_name)

        # Headers
        headers = ["CVE ID", "Severity", "CVSS v3", "CVSS v2", "Published",
                   "CWE", "Description", "References"]

        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for row_idx, cve in enumerate(cves, start=2):
            sheet.cell(row=row_idx, column=1, value=cve.get("CVE ID"))

            severity_cell = sheet.cell(row=row_idx, column=2, value=cve.get("CVSS v3 Severity"))
            severity = cve.get("CVSS v3 Severity", "")
            if severity == "CRITICAL":
                severity_cell.font = Font(color="C00000", bold=True)
            elif severity == "HIGH":
                severity_cell.font = Font(color="FF6600", bold=True)
            elif severity == "MEDIUM":
                severity_cell.font = Font(color="FFC000")

            sheet.cell(row=row_idx, column=3, value=cve.get("CVSS v3 Score"))
            sheet.cell(row=row_idx, column=4, value=cve.get("CVSS v2 Score"))
            sheet.cell(row=row_idx, column=5, value=cve.get("Published Date"))
            sheet.cell(row=row_idx, column=6, value=cve.get("CWE"))
            sheet.cell(row=row_idx, column=7, value=cve.get("Description"))
            sheet.cell(row=row_idx, column=8, value=cve.get("References"))

            # Wrap text
            for col in [7, 8]:
                sheet.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")

        # Column widths
        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 12
        sheet.column_dimensions["C"].width = 10
        sheet.column_dimensions["D"].width = 10
        sheet.column_dimensions["E"].width = 12
        sheet.column_dimensions["F"].width = 25
        sheet.column_dimensions["G"].width = 70
        sheet.column_dimensions["H"].width = 50

        sheet.freeze_panes = "A2"

    wb.save(output_file)
    print(f"\n{'='*70}")
    print(f"✓ Excel report saved: {output_file}")
    print(f"{'='*70}")

def main():
    print("\n" + "="*70)
    print("COMPREHENSIVE REGEX ENGINE CVE REPORT GENERATOR")
    print("Data Source: NIST National Vulnerability Database (NVD) API v2.0")
    print("="*70)

    cve_data_by_engine = {}

    for engine_name, engine_config in REGEX_ENGINES.items():
        cves = collect_cves_for_engine(engine_name, engine_config)
        cve_data_by_engine[engine_name] = cves

    # Handle shared PCRE/PCRE2 CVEs
    print(f"\n{'='*70}")
    print("Checking for shared PCRE/PCRE2 CVEs...")
    print(f"{'='*70}")
    add_shared_pcre_cves(cve_data_by_engine)

    output_file = "/mnt/user-data/outputs/regex_cve_report.xlsx"
    create_excel_report(cve_data_by_engine, output_file)

    # Summary
    print("\nFINAL SUMMARY:")
    print("-" * 70)
    total_unique = len(set(
        cve["CVE ID"]
        for cves in cve_data_by_engine.values()
        for cve in cves
    ))
    total_entries = sum(len(cves) for cves in cve_data_by_engine.values())

    for engine_name in sorted(cve_data_by_engine.keys()):
        count = len(cve_data_by_engine[engine_name])
        if count > 0:
            print(f"  {engine_name:.<40} {count:>4} CVEs")
    print("-" * 70)
    print(f"  {'TOTAL ENTRIES (with duplicates)':.<40} {total_entries:>4}")
    print(f"  {'UNIQUE CVEs':.<40} {total_unique:>4}")
    print("="*70)

if __name__ == "__main__":
    main()
